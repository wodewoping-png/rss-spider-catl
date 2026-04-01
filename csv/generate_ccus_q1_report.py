from collections import OrderedDict
from pathlib import Path
import os

os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib")

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
NEWS_CSV = BASE_DIR / "carbonherald_q1_2026_news.csv"
REPORT_XLSX = BASE_DIR / "ccus_q1_2026_中文分析报告.xlsx"
REPORT_MD = BASE_DIR / "ccus_q1_2026_brief_report.md"

MAJOR_BAR_PNG = BASE_DIR / "ccus_q1_2026_major_category_bar.png"
MAJOR_PIE_PNG = BASE_DIR / "ccus_q1_2026_major_category_pie.png"
REMOVAL_BAR_PNG = BASE_DIR / "ccus_q1_2026_removal_tech_bar.png"
REMOVAL_PIE_PNG = BASE_DIR / "ccus_q1_2026_removal_tech_pie.png"
MONTHLY_TREND_PNG = BASE_DIR / "ccus_q1_2026_monthly_trend.png"

LABEL_ZH = {
    "Removal": "碳移除",
    "Capture": "碳捕集",
    "Storage": "碳封存",
    "Biomass": "生物质",
    "Markets": "碳市场",
    "Direct Air Capture": "直接空气捕集",
    "Mineralization": "矿化",
    "Policy": "政策",
    "Forests": "森林碳汇",
    "Farming": "农业碳汇",
    "Utilization": "碳利用",
    "BECCS": "生物质能碳捕集与封存",
    "Ocean": "海洋碳汇",
}

LABEL_EN = {
    "碳封存": "Storage",
    "碳捕集": "Capture",
    "碳移除": "Removal",
    "碳利用": "Utilization",
    "生物质能碳捕集与封存": "BECCS",
    "海洋碳汇": "Ocean CDR",
    "农业碳汇": "Ag CDR",
    "森林碳汇": "Forest CDR",
    "矿化": "Mineralization",
    "直接空气捕集": "DAC",
    "生物质": "Biomass",
}

MAJOR_CATEGORY_MAP = OrderedDict(
    [
        ("碳封存", ["Storage"]),
        ("碳捕集", ["Capture"]),
        ("碳移除", ["Removal", "BECCS", "Ocean", "Farming", "Forests", "Mineralization", "Direct Air Capture", "Biomass"]),
        ("碳利用", ["Utilization"]),
    ]
)

REMOVAL_TECH_MAP = OrderedDict(
    [
        ("生物质能碳捕集与封存", "BECCS"),
        ("海洋碳汇", "Ocean"),
        ("农业碳汇", "Farming"),
        ("森林碳汇", "Forests"),
        ("矿化", "Mineralization"),
        ("直接空气捕集", "Direct Air Capture"),
        ("生物质", "Biomass"),
    ]
)

LABEL_TO_MAJOR = {
    label: category for category, labels in MAJOR_CATEGORY_MAP.items() for label in labels
}

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="DCE6F1")
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
BODY_FONT = Font(name="Microsoft YaHei", size=10)


def to_zh_label(label: str) -> str:
    return LABEL_ZH.get(label, label)


def count_labels(news: pd.DataFrame) -> pd.Series:
    return (
        news["labels"]
        .fillna("")
        .str.split("|")
        .explode()
        .replace("", pd.NA)
        .dropna()
        .value_counts()
    )


def build_major_counts(label_counts: pd.Series) -> pd.DataFrame:
    rows = []
    total = sum(label_counts.get(label, 0) for labels in MAJOR_CATEGORY_MAP.values() for label in labels)
    for category, labels in MAJOR_CATEGORY_MAP.items():
        count = int(sum(label_counts.get(label, 0) for label in labels))
        rows.append(
            {
                "category_zh": category,
                "source_labels": "|".join(labels),
                "news_count": count,
                "share_pct": count / total if total else 0,
            }
        )
    return pd.DataFrame(rows)


def build_removal_counts(label_counts: pd.Series) -> pd.DataFrame:
    rows = []
    total = sum(label_counts.get(label, 0) for label in REMOVAL_TECH_MAP.values())
    for tech_zh, label in REMOVAL_TECH_MAP.items():
        count = int(label_counts.get(label, 0))
        rows.append(
            {
                "tech_zh": tech_zh,
                "source_label": label,
                "news_count": count,
                "share_pct": count / total if total else 0,
            }
        )
    return pd.DataFrame(rows)


def build_monthly_major_dist(news: pd.DataFrame) -> pd.DataFrame:
    exploded = (
        news.assign(month=news["date"].dt.strftime("%Y-%m"))
        .assign(label=news["labels"].fillna("").str.split("|"))
        .explode("label")
        .replace({"label": {"": pd.NA}})
        .dropna(subset=["label"])
        .assign(major=lambda df: df["label"].map(LABEL_TO_MAJOR))
        .dropna(subset=["major"])
    )
    monthly_major_dist = (
        exploded.groupby(["major", "month"])
        .size()
        .unstack(fill_value=0)
        .reindex(list(MAJOR_CATEGORY_MAP.keys()), fill_value=0)
    )
    return monthly_major_dist.reindex(sorted(monthly_major_dist.columns), axis=1, fill_value=0)


def build_top_major_by_month(monthly_major_dist: pd.DataFrame) -> dict[str, pd.Series]:
    result = {}
    for month in monthly_major_dist.columns:
        result[month] = monthly_major_dist[month].sort_values(ascending=False)
    return result


def prepare_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.Series, pd.DataFrame, dict[str, pd.Series]]:
    news = pd.read_csv(NEWS_CSV)
    news["date"] = pd.to_datetime(news["date"])

    label_counts = count_labels(news)
    major_counts = build_major_counts(label_counts)
    removal_counts = build_removal_counts(label_counts)
    monthly_counts = news.groupby(news["date"].dt.strftime("%Y-%m")).size().sort_index()
    monthly_major_dist = build_monthly_major_dist(news)
    top_major_by_month = build_top_major_by_month(monthly_major_dist)

    return news, major_counts, removal_counts, monthly_counts, monthly_major_dist, top_major_by_month


def build_summary_lines(
    news: pd.DataFrame,
    major_counts: pd.DataFrame,
    removal_counts: pd.DataFrame,
    monthly_counts: pd.Series,
    monthly_major_dist: pd.DataFrame,
    top_major_by_month: dict[str, pd.Series],
) -> list[str]:
    total_news = len(news)
    total_major_mentions = int(major_counts["news_count"].sum())
    total_removal_mentions = int(removal_counts["news_count"].sum())

    major_rank = major_counts.sort_values("news_count", ascending=False).reset_index(drop=True)
    removal_rank = removal_counts.sort_values("news_count", ascending=False).reset_index(drop=True)

    max_month = monthly_counts.idxmax()
    min_month = monthly_counts.idxmin()
    monthly_delta = monthly_counts.diff()

    removal_by_month = [int(monthly_major_dist.loc["碳移除", month]) for month in monthly_counts.index]

    lines = [
        "一、数据范围",
        f"新闻明细文件：{NEWS_CSV.name}",
        "统计区间：2026-01-01 至 2026-03-31",
        f"新闻总量：{total_news} 篇",
        "本次分析仅保留四个大类：碳封存、碳捕集、碳移除、碳利用。",
        "不纳入统计：碳市场、政策及其他非上述四类标签。",
        f"四大类合计提及量：{total_major_mentions} 次",
        f"碳移除细分技术合计提及量：{total_removal_mentions} 次",
        "",
        "二、核心结论",
        (
            f"一季度新闻量总体稳定，1月 {int(monthly_counts['2026-01'])} 篇，"
            f"2月 {int(monthly_counts['2026-02'])} 篇，3月 {int(monthly_counts['2026-03'])} 篇，"
            f"峰值出现在 {max_month}，低点出现在 {min_month}。"
        ),
        (
            f"四大类中，“{major_rank.iloc[0]['category_zh']}”以 {int(major_rank.iloc[0]['news_count'])} 次居首，"
            f"占四大类提及量的 {major_rank.iloc[0]['share_pct'] * 100:.1f}%；"
            f"“{major_rank.iloc[1]['category_zh']}”以 {int(major_rank.iloc[1]['news_count'])} 次位列第二。"
        ),
        (
            f"“碳移除”连续三个月保持第一，分别为 {removal_by_month[0]}、{removal_by_month[1]}、"
            f"{removal_by_month[2]} 次，说明移除相关议题仍是季度主线。"
        ),
        (
            f"“碳捕集”在 2026-02 达到季度内月度高点 {int(monthly_major_dist.loc['碳捕集', '2026-02'])} 次，"
            f"随后 2026-03 回落至 {int(monthly_major_dist.loc['碳捕集', '2026-03'])} 次。"
        ),
        (
            f"碳移除细分技术中，“{removal_rank.iloc[0]['tech_zh']}”提及 {int(removal_rank.iloc[0]['news_count'])} 次居首，"
            f"“{removal_rank.iloc[1]['tech_zh']}”和“{removal_rank.iloc[2]['tech_zh']}”分别为 "
            f"{int(removal_rank.iloc[1]['news_count'])} 次和 {int(removal_rank.iloc[2]['news_count'])} 次。"
        ),
        "",
        "三、月度走势",
        f"2026-01：{int(monthly_counts['2026-01'])} 篇，作为基准月。",
        f"2026-02：{int(monthly_counts['2026-02'])} 篇，较 1 月变动 {int(monthly_delta['2026-02']):+d} 篇。",
        f"2026-03：{int(monthly_counts['2026-03'])} 篇，较 2 月变动 {int(monthly_delta['2026-03']):+d} 篇。",
        "",
        "各月四大类排序：",
    ]

    for month in monthly_counts.index:
        counts = top_major_by_month[month]
        summary = "，".join(f"{label} {int(value)} 次" for label, value in counts.items())
        lines.append(f"{month}：{summary}")

    lines.extend(
        [
            "",
            "四、简要判断",
            "四大类口径下，季度关注度明显向碳移除集中，说明媒体仍主要围绕负排放、碳移除项目和相关技术商业化进展展开报道。",
            "碳捕集和碳封存在项目签约、工程建设和基础设施议题上保持稳定存在，但总量明显低于碳移除，反映产业链关注点仍偏前端技术和新项目动态。",
            "在碳移除内部，生物质、直接空气捕集和矿化形成第一梯队，森林碳汇、农业碳汇、海洋碳汇和 BECCS 构成第二梯队，显示技术路线分布较广，但热度仍向工程化程度更高的方向集中。",
        ]
    )
    return lines


def build_md_report(
    news: pd.DataFrame,
    major_counts: pd.DataFrame,
    removal_counts: pd.DataFrame,
    monthly_counts: pd.Series,
    top_major_by_month: dict[str, pd.Series],
) -> str:
    lines = [
        "# CCUS 2026年一季度趋势简报",
        "",
        "## 1. 数据范围",
        f"- 新闻明细文件：`{NEWS_CSV.name}`",
        "- 统计区间：2026-01-01 至 2026-03-31",
        f"- 新闻总量：{len(news)} 篇",
        "- 分析口径：仅保留碳封存、碳捕集、碳移除、碳利用四个大类",
        "- 不纳入统计：碳市场、政策及其他非上述四类标签",
        f"- 四大类合计提及量：{int(major_counts['news_count'].sum())} 次",
        f"- 碳移除细分技术合计提及量：{int(removal_counts['news_count'].sum())} 次",
        "",
        "## 2. 核心结论",
    ]

    major_rank = major_counts.sort_values("news_count", ascending=False).reset_index(drop=True)
    removal_rank = removal_counts.sort_values("news_count", ascending=False).reset_index(drop=True)

    lines.extend(
        [
            (
                f"- 一季度新闻量总体稳定，1月 {int(monthly_counts['2026-01'])} 篇，2月 {int(monthly_counts['2026-02'])} 篇，"
                f"3月 {int(monthly_counts['2026-03'])} 篇，峰值出现在 {monthly_counts.idxmax()}，低点出现在 {monthly_counts.idxmin()}。"
            ),
            (
                f"- 四大类中 `{major_rank.iloc[0]['category_zh']}` 居首，为 {int(major_rank.iloc[0]['news_count'])} 次，"
                f"占四大类提及量的 {major_rank.iloc[0]['share_pct'] * 100:.1f}%；`{major_rank.iloc[1]['category_zh']}` 以 "
                f"{int(major_rank.iloc[1]['news_count'])} 次位列第二。"
            ),
            (
                f"- `碳移除` 连续三个月保持第一，分别为 "
                f"{int(top_major_by_month['2026-01']['碳移除'])}、{int(top_major_by_month['2026-02']['碳移除'])}、"
                f"{int(top_major_by_month['2026-03']['碳移除'])} 次。"
            ),
            (
                f"- `碳捕集` 在 2 月达到季度内月度高点 {int(top_major_by_month['2026-02']['碳捕集'])} 次，"
                f"随后 3 月回落至 {int(top_major_by_month['2026-03']['碳捕集'])} 次。"
            ),
            (
                f"- 碳移除细分技术中，`{removal_rank.iloc[0]['tech_zh']}` 以 {int(removal_rank.iloc[0]['news_count'])} 次居首，"
                f"`{removal_rank.iloc[1]['tech_zh']}` 和 `{removal_rank.iloc[2]['tech_zh']}` 分别为 "
                f"{int(removal_rank.iloc[1]['news_count'])} 次和 {int(removal_rank.iloc[2]['news_count'])} 次。"
            ),
            "",
            "## 3. 四大类分布",
        ]
    )

    for row in major_counts.itertuples(index=False):
        lines.append(f"- {row.category_zh}: {int(row.news_count)} 次，占比 {row.share_pct * 100:.1f}%")

    lines.extend(["", "## 4. 碳移除技术分布"])
    for row in removal_counts.sort_values("news_count", ascending=False).itertuples(index=False):
        lines.append(f"- {row.tech_zh}: {int(row.news_count)} 次，占比 {row.share_pct * 100:.1f}%")

    lines.extend(
        [
            "",
            "## 5. 月度走势",
            f"- 2026-01: {int(monthly_counts['2026-01'])} 篇，环比基准月。",
            f"- 2026-02: {int(monthly_counts['2026-02'])} 篇，较 1 月变动 {int(monthly_counts.diff()['2026-02']):+d} 篇。",
            f"- 2026-03: {int(monthly_counts['2026-03'])} 篇，较 2 月变动 {int(monthly_counts.diff()['2026-03']):+d} 篇。",
            "",
            "各月四大类排序：",
        ]
    )

    for month in monthly_counts.index:
        counts = top_major_by_month[month]
        summary = ", ".join(f"{label} {int(value)}" for label, value in counts.items())
        lines.append(f"- {month}: {summary}")

    lines.extend(
        [
            "",
            "## 6. 图表文件",
            f"- 四大类条形图：`{MAJOR_BAR_PNG.name}`",
            f"- 四大类饼形图：`{MAJOR_PIE_PNG.name}`",
            f"- 碳移除技术条形图：`{REMOVAL_BAR_PNG.name}`",
            f"- 碳移除技术饼形图：`{REMOVAL_PIE_PNG.name}`",
            f"- 月度趋势图：`{MONTHLY_TREND_PNG.name}`",
            "",
            "## 7. 简要判断",
            "- 四大类口径下，季度关注度明显向碳移除集中，说明媒体仍主要围绕负排放、碳移除项目和相关技术商业化进展展开报道。",
            "- 碳捕集和碳封存在项目签约、工程建设和基础设施议题上保持稳定存在，但总量明显低于碳移除。",
            "- 在碳移除内部，生物质、直接空气捕集和矿化形成第一梯队，其余路线共同构成较分散的第二梯队。",
        ]
    )
    return "\n".join(lines) + "\n"


def style_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_summary_sheet(ws, summary_lines: list[str]) -> None:
    ws.title = "摘要说明"
    ws.merge_cells("A1:D1")
    ws["A1"] = "CCUS 2026年一季度中文分析报告"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    row = 3
    for line in summary_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=line)
        if line and line[1:2] == "、":
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        else:
            cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 28


def add_bar_chart(ws, title: str, category_title: str, value_title: str, category_ref, data_ref, anchor: str) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = category_title
    chart.x_axis.title = value_title
    chart.height = 8
    chart.width = 16
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(category_ref)
    chart.legend = None
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    ws.add_chart(chart, anchor)


def add_pie_chart(ws, title: str, category_ref, data_ref, anchor: str) -> None:
    chart = PieChart()
    chart.style = 10
    chart.title = title
    chart.height = 10
    chart.width = 12
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(category_ref)
    chart.dLbls = DataLabelList()
    chart.dLbls.showPercent = True
    chart.dLbls.showLeaderLines = True
    chart.dLbls.showLegendKey = False
    ws.add_chart(chart, anchor)


def write_major_sheet(ws, major_counts: pd.DataFrame) -> None:
    ws.title = "四大类统计"
    headers = ["分类", "对应标签", "提及次数", "占比"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    display_df = major_counts.sort_values("news_count", ascending=False).reset_index(drop=True)
    for idx, row in enumerate(display_df.itertuples(index=False), start=2):
        ws.cell(row=idx, column=1, value=row.category_zh)
        ws.cell(row=idx, column=2, value=row.source_labels)
        ws.cell(row=idx, column=3, value=int(row.news_count))
        pct_cell = ws.cell(row=idx, column=4, value=float(row.share_pct))
        pct_cell.number_format = "0.0%"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A2"

    max_row = len(display_df) + 1
    category_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    data_ref = Reference(ws, min_col=3, min_row=1, max_row=max_row)
    pie_ref = Reference(ws, min_col=4, min_row=1, max_row=max_row)

    add_bar_chart(ws, "CCUS 四大类提及次数", "分类", "提及次数", category_ref, data_ref, "F2")
    add_pie_chart(ws, "CCUS 四大类占比", category_ref, pie_ref, "F20")


def write_removal_sheet(ws, removal_counts: pd.DataFrame) -> None:
    ws.title = "碳移除技术"
    headers = ["技术方向", "对应标签", "提及次数", "占比"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    display_df = removal_counts.sort_values("news_count", ascending=False).reset_index(drop=True)
    for idx, row in enumerate(display_df.itertuples(index=False), start=2):
        ws.cell(row=idx, column=1, value=row.tech_zh)
        ws.cell(row=idx, column=2, value=row.source_label)
        ws.cell(row=idx, column=3, value=int(row.news_count))
        pct_cell = ws.cell(row=idx, column=4, value=float(row.share_pct))
        pct_cell.number_format = "0.0%"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A2"

    max_row = len(display_df) + 1
    category_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    data_ref = Reference(ws, min_col=3, min_row=1, max_row=max_row)
    pie_ref = Reference(ws, min_col=4, min_row=1, max_row=max_row)

    add_bar_chart(ws, "碳移除细分技术提及次数", "技术方向", "提及次数", category_ref, data_ref, "F2")
    add_pie_chart(ws, "碳移除细分技术占比", category_ref, pie_ref, "F22")


def write_monthly_sheet(
    ws,
    monthly_counts: pd.Series,
    top_major_by_month: dict[str, pd.Series],
    monthly_major_dist: pd.DataFrame,
) -> None:
    ws.title = "月度趋势"
    headers = ["月份", "新闻数量", "四大类排序摘要"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    months = list(monthly_counts.index)
    for idx, month in enumerate(months, start=2):
        ws.cell(row=idx, column=1, value=month)
        ws.cell(row=idx, column=2, value=int(monthly_counts[month]))
        summary = "，".join(f"{label} {int(value)} 次" for label, value in top_major_by_month[month].items())
        ws.cell(row=idx, column=3, value=summary)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 46
    ws.freeze_panes = "A2"

    line_chart = LineChart()
    line_chart.style = 10
    line_chart.title = "CCUS 月度新闻走势"
    line_chart.y_axis.title = "新闻数量"
    line_chart.x_axis.title = "月份"
    line_chart.height = 8
    line_chart.width = 16
    line_chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=len(months) + 1), titles_from_data=True)
    line_chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(months) + 1))
    line_chart.dLbls = DataLabelList()
    line_chart.dLbls.showVal = True
    ws.add_chart(line_chart, "E2")

    start_row = 8
    start_col = 5
    ws.cell(row=start_row, column=start_col, value="四大类月度分布数据").font = SECTION_FONT
    ws.cell(row=start_row, column=start_col).fill = SECTION_FILL

    ws.cell(row=start_row + 1, column=start_col, value="分类").font = SECTION_FONT
    ws.cell(row=start_row + 1, column=start_col).fill = HEADER_FILL
    for offset, month in enumerate(monthly_major_dist.columns, start=1):
        cell = ws.cell(row=start_row + 1, column=start_col + offset, value=month)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    total_header = ws.cell(
        row=start_row + 1,
        column=start_col + len(monthly_major_dist.columns) + 1,
        value="Q1合计",
    )
    total_header.font = SECTION_FONT
    total_header.fill = HEADER_FILL
    total_header.alignment = Alignment(horizontal="center")

    for row_offset, (label, values) in enumerate(monthly_major_dist.iterrows(), start=2):
        ws.cell(row=start_row + row_offset, column=start_col, value=label)
        for col_offset, value in enumerate(values.tolist(), start=1):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=int(value))
        ws.cell(
            row=start_row + row_offset,
            column=start_col + len(monthly_major_dist.columns) + 1,
            value=int(sum(values.tolist())),
        )

    dist_chart = BarChart()
    dist_chart.type = "bar"
    dist_chart.grouping = "stacked"
    dist_chart.overlap = 100
    dist_chart.style = 10
    dist_chart.title = "四大类月度分布"
    dist_chart.x_axis.title = "提及次数"
    dist_chart.y_axis.title = "分类"
    dist_chart.height = 10
    dist_chart.width = 18
    data_min_row = start_row + 1
    data_max_row = start_row + 1 + len(monthly_major_dist)
    data_max_col = start_col + len(monthly_major_dist.columns)
    dist_chart.add_data(
        Reference(ws, min_col=start_col + 1, min_row=data_min_row, max_col=data_max_col, max_row=data_max_row),
        titles_from_data=True,
    )
    dist_chart.set_categories(Reference(ws, min_col=start_col, min_row=start_row + 2, max_row=data_max_row))
    dist_chart.legend.position = "r"
    ws.add_chart(dist_chart, "J2")


def write_news_sheet(ws, news: pd.DataFrame) -> None:
    ws.title = "新闻明细"
    output = news.copy()
    output["labels_zh"] = (
        output["labels"]
        .fillna("")
        .str.split("|")
        .apply(lambda items: "|".join(to_zh_label(item) for item in items if item))
    )
    output["major_categories_zh"] = (
        output["labels"]
        .fillna("")
        .str.split("|")
        .apply(
            lambda items: "|".join(
                dict.fromkeys(
                    LABEL_TO_MAJOR[item]
                    for item in items
                    if item in LABEL_TO_MAJOR
                )
            )
        )
    )

    headers = ["日期", "标题", "原始标签", "中文标签", "四大类归属", "链接"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = SECTION_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(output.itertuples(index=False), start=2):
        ws.cell(row=row_idx, column=1, value=row.date.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=2, value=row.title)
        ws.cell(row=row_idx, column=3, value=row.labels)
        ws.cell(row=row_idx, column=4, value=row.labels_zh)
        ws.cell(row=row_idx, column=5, value=row.major_categories_zh)
        ws.cell(row=row_idx, column=6, value=row.url)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 80
    ws.freeze_panes = "A2"


def build_workbook(
    news: pd.DataFrame,
    major_counts: pd.DataFrame,
    removal_counts: pd.DataFrame,
    monthly_counts: pd.Series,
    monthly_major_dist: pd.DataFrame,
    top_major_by_month: dict[str, pd.Series],
) -> Workbook:
    wb = Workbook()
    summary_ws = wb.active

    summary_lines = build_summary_lines(
        news,
        major_counts,
        removal_counts,
        monthly_counts,
        monthly_major_dist,
        top_major_by_month,
    )
    write_summary_sheet(summary_ws, summary_lines)
    write_major_sheet(wb.create_sheet(), major_counts)
    write_removal_sheet(wb.create_sheet(), removal_counts)
    write_monthly_sheet(wb.create_sheet(), monthly_counts, top_major_by_month, monthly_major_dist)
    write_news_sheet(wb.create_sheet(), news)

    for ws in wb.worksheets:
        style_sheet(ws)

    return wb


def configure_matplotlib() -> None:
    plt.rcParams["font.sans-serif"] = [
        "Noto Sans CJK SC",
        "Microsoft YaHei",
        "SimHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False


def save_bar_chart(df: pd.DataFrame, label_col: str, value_col: str, title: str, output_path: Path, color: str) -> None:
    configure_matplotlib()
    display_df = df.sort_values(value_col, ascending=True)
    fig, ax = plt.subplots(figsize=(10, 6))
    bars = ax.barh(display_df[label_col], display_df[value_col], color=color)
    ax.set_title(title)
    ax.set_xlabel("Mentions")
    ax.set_ylabel("")
    ax.grid(axis="x", linestyle="--", alpha=0.3)
    ax.set_axisbelow(True)
    for bar, value in zip(bars, display_df[value_col]):
        ax.text(value + max(display_df[value_col]) * 0.01, bar.get_y() + bar.get_height() / 2, f"{int(value)}", va="center")
    fig.tight_layout()
    fig.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def save_pie_chart(df: pd.DataFrame, label_col: str, value_col: str, title: str, output_path: Path, colors: list[str]) -> None:
    configure_matplotlib()
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.pie(
        df[value_col],
        labels=df[label_col],
        autopct="%1.1f%%",
        startangle=90,
        colors=colors[: len(df)],
        pctdistance=0.75,
    )
    centre = plt.Circle((0, 0), 0.45, fc="white")
    fig.gca().add_artist(centre)
    ax.set_title(title)
    fig.tight_layout()
    fig.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def save_monthly_trend_chart(monthly_counts: pd.Series, output_path: Path) -> None:
    configure_matplotlib()
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.plot(monthly_counts.index.tolist(), monthly_counts.tolist(), marker="o", linewidth=2.5, color="#2F6B7C")
    ax.set_title("CCUS Q1 2026 Monthly News Trend")
    ax.set_xlabel("Month")
    ax.set_ylabel("News Count")
    ax.grid(axis="y", linestyle="--", alpha=0.3)
    for x, y in zip(monthly_counts.index.tolist(), monthly_counts.tolist()):
        ax.text(x, y + 1, f"{int(y)}", ha="center")
    fig.tight_layout()
    fig.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def generate_png_charts(major_counts: pd.DataFrame, removal_counts: pd.DataFrame, monthly_counts: pd.Series) -> None:
    major_chart_df = major_counts.sort_values("news_count", ascending=False).assign(
        category_en=lambda df: df["category_zh"].map(LABEL_EN)
    )
    removal_chart_df = removal_counts.sort_values("news_count", ascending=False).assign(
        tech_en=lambda df: df["tech_zh"].map(LABEL_EN)
    )

    save_bar_chart(major_chart_df, "category_en", "news_count", "Q1 2026 CCUS Major Categories", MAJOR_BAR_PNG, "#2F6B7C")
    save_pie_chart(
        major_chart_df,
        "category_en",
        "news_count",
        "Q1 2026 CCUS Major Category Share",
        MAJOR_PIE_PNG,
        ["#2F6B7C", "#5E8C61", "#D9A441", "#B85C38"],
    )
    save_bar_chart(removal_chart_df, "tech_en", "news_count", "Q1 2026 Carbon Removal Technologies", REMOVAL_BAR_PNG, "#5E8C61")
    save_pie_chart(
        removal_chart_df,
        "tech_en",
        "news_count",
        "Q1 2026 Carbon Removal Tech Share",
        REMOVAL_PIE_PNG,
        ["#5E8C61", "#7AA874", "#97C17C", "#B9D48B", "#D9A441", "#C97D60", "#8A5082"],
    )
    save_monthly_trend_chart(monthly_counts, MONTHLY_TREND_PNG)


def main() -> None:
    news, major_counts, removal_counts, monthly_counts, monthly_major_dist, top_major_by_month = prepare_data()

    workbook = build_workbook(
        news,
        major_counts,
        removal_counts,
        monthly_counts,
        monthly_major_dist,
        top_major_by_month,
    )
    workbook.save(REPORT_XLSX)

    generate_png_charts(major_counts, removal_counts, monthly_counts)
    REPORT_MD.write_text(
        build_md_report(news, major_counts, removal_counts, monthly_counts, top_major_by_month),
        encoding="utf-8",
    )

    print(f"Generated: {REPORT_XLSX.name}")
    print(f"Updated: {REPORT_MD.name}")
    print(f"Generated charts: {MAJOR_BAR_PNG.name}, {MAJOR_PIE_PNG.name}, {REMOVAL_BAR_PNG.name}, {REMOVAL_PIE_PNG.name}, {MONTHLY_TREND_PNG.name}")


if __name__ == "__main__":
    main()
