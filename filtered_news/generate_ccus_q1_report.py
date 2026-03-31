from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
RECLASSIFIED_PATH = BASE_DIR / "all_literature_reclassified_11cats.xlsx"
CLASSIFIED_PATH = BASE_DIR / "all_literature_classified.xlsx"
OUTPUT_XLSX = BASE_DIR / "2026_CCUS_Q1_文献分析报告.xlsx"
OUTPUT_MD = BASE_DIR / "2026_CCUS_Q1_文献分析简报.md"


def load_q1_dataset() -> pd.DataFrame:
    new_sheets = pd.read_excel(RECLASSIFIED_PATH, sheet_name=None)
    new_df = pd.concat(
        [sheet_df.assign(category=sheet_name) for sheet_name, sheet_df in new_sheets.items() if sheet_name != "其他"],
        ignore_index=True,
    )

    old_sheets = pd.read_excel(CLASSIFIED_PATH, sheet_name=None)
    old_df = pd.concat(old_sheets.values(), ignore_index=True)

    for df in (new_df, old_df):
        df["merge_key"] = df["title"].astype(str).str.strip().str.lower()

    merged = new_df.merge(
        old_df[["merge_key", "pub_date", "source"]],
        on="merge_key",
        how="left",
        suffixes=("_new", "_old"),
    )

    merged["pub_date_old"] = pd.to_datetime(merged["pub_date_old"], utc=True, errors="coerce")
    q1 = merged[
        (merged["pub_date_old"] >= pd.Timestamp("2026-01-01", tz="UTC"))
        & (merged["pub_date_old"] < pd.Timestamp("2026-04-01", tz="UTC"))
    ].copy()
    q1["month"] = q1["pub_date_old"].dt.strftime("%Y-%m")
    q1["pub_date_str"] = q1["pub_date_old"].dt.strftime("%Y-%m-%d")
    return q1


def build_summary_tables(q1: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    category_counts = (
        q1["category"]
        .value_counts()
        .rename_axis("category")
        .reset_index(name="count")
        .sort_values(["count", "category"], ascending=[False, True])
        .reset_index(drop=True)
    )
    total = int(category_counts["count"].sum())
    category_counts["share"] = category_counts["count"] / total

    month_order = ["2026-01", "2026-02", "2026-03"]
    monthly_trend = (
        pd.crosstab(q1["category"], q1["month"])
        .reindex(columns=month_order, fill_value=0)
        .reindex(category_counts["category"], fill_value=0)
        .reset_index()
    )
    monthly_trend["Q1_total"] = monthly_trend[month_order].sum(axis=1)

    month_total = (
        q1["month"].value_counts().reindex(month_order, fill_value=0).rename_axis("month").reset_index(name="count")
    )
    month_total["share"] = month_total["count"] / total
    return category_counts, monthly_trend, month_total


def generate_insights(category_counts: pd.DataFrame, monthly_trend: pd.DataFrame, month_total: pd.DataFrame) -> list[str]:
    total = int(category_counts["count"].sum())
    top3 = category_counts.head(3)
    top3_share = top3["share"].sum()

    month_values = month_total.set_index("month")["count"].to_dict()
    jan = int(month_values.get("2026-01", 0))
    feb = int(month_values.get("2026-02", 0))
    mar = int(month_values.get("2026-03", 0))

    trend_map = monthly_trend.set_index("category")
    elec = trend_map.loc["CO₂电转化", ["2026-01", "2026-02", "2026-03"]].tolist()
    photo = trend_map.loc["CO₂光转化", ["2026-01", "2026-02", "2026-03"]].tolist()
    capture = trend_map.loc["CO₂捕集与分离", ["2026-01", "2026-02", "2026-03"]].tolist()
    policy = trend_map.loc["政策与产业化", ["2026-01", "2026-02", "2026-03"]].tolist()

    return [
        f"2026年Q1样本共纳入 {total} 篇 CCUS 文献，统计范围排除了“其他”分类。",
        (
            f"文献最集中的三大领域为 {top3.iloc[0]['category']}、{top3.iloc[1]['category']}、"
            f"{top3.iloc[2]['category']}，合计占比 {top3_share:.1%}，说明研究热点仍集中在转化端与捕集分离端。"
        ),
        f"季度总量从 1 月 {jan} 篇下降到 2 月 {feb} 篇，再到 3 月 {mar} 篇，样本呈逐月回落。",
        (
            f"CO₂电转化在三个月分别为 {elec[0]}/{elec[1]}/{elec[2]} 篇，持续保持第一；"
            f"CO₂光转化为 {photo[0]}/{photo[1]}/{photo[2]} 篇，3 月明显收缩。"
        ),
        (
            f"CO₂捕集与分离在三个月分别为 {capture[0]}/{capture[1]}/{capture[2]} 篇，整体相对稳健；"
            f"政策与产业化在 2 月达到 {policy[1]} 篇，较 1 月和 3 月更活跃。"
        ),
        "矿化、生物转化、运输与封存、系统集成与工艺耦合等方向样本量较小，更适合结合后续季度持续跟踪，而不宜只看单月波动做强结论。",
        "说明：季度趋势基于当前工作区样本统计，文献日期来自 `all_literature_classified.xlsx` 中与 11 分类总表按标题回溯匹配的 `pub_date`。",
    ]


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 42)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_df(ws, df: pd.DataFrame, percentage_columns: set[str] | None = None) -> None:
    percentage_columns = percentage_columns or set()
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    style_sheet(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
            if ws.cell(1, cell.column).value in percentage_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"


def build_excel_report(q1: pd.DataFrame, category_counts: pd.DataFrame, monthly_trend: pd.DataFrame, month_total: pd.DataFrame) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "执行摘要"
    summary_ws["A1"] = "2026 CCUS Q1 文献分析报告"
    summary_ws["A1"].font = Font(size=16, bold=True)
    summary_ws["A3"] = "统计口径"
    summary_ws["A3"].font = Font(bold=True)
    summary_ws["B3"] = "排除“其他”分类；按 2026-01-01 至 2026-03-31 的文献日期统计。"
    summary_ws["A4"] = "样本总量"
    summary_ws["B4"] = int(len(q1))

    insights = generate_insights(category_counts, monthly_trend, month_total)
    summary_ws["A6"] = "核心结论"
    summary_ws["A6"].font = Font(bold=True)
    for idx, text in enumerate(insights, start=7):
        summary_ws[f"A{idx}"] = f"{idx - 6}. {text}"

    summary_ws.column_dimensions["A"].width = 16
    summary_ws.column_dimensions["B"].width = 96

    counts_ws = wb.create_sheet("分类统计")
    counts_export = category_counts.copy()
    counts_export.columns = ["分类", "文献数", "占比"]
    write_df(counts_ws, counts_export, percentage_columns={"占比"})

    trend_ws = wb.create_sheet("月度趋势")
    trend_export = monthly_trend.copy()
    trend_export.columns = ["分类", "2026-01", "2026-02", "2026-03", "Q1合计"]
    write_df(trend_ws, trend_export)

    month_ws = wb.create_sheet("月度总量")
    month_export = month_total.copy()
    month_export.columns = ["月份", "文献数", "占比"]
    write_df(month_ws, month_export, percentage_columns={"占比"})

    detail_ws = wb.create_sheet("Q1样本明细")
    detail_cols = ["title", "category", "pub_date_str", "source_old", "link"]
    detail_export = q1[detail_cols].copy().sort_values(["pub_date_str", "category", "title"])
    detail_export.columns = ["标题", "分类", "日期", "来源", "链接"]
    write_df(detail_ws, detail_export)

    bar_chart = BarChart()
    bar_chart.title = "2026 Q1 各领域文献数"
    bar_chart.y_axis.title = "文献数"
    bar_chart.x_axis.title = "领域"
    data = Reference(counts_ws, min_col=2, min_row=1, max_row=counts_ws.max_row)
    labels = Reference(counts_ws, min_col=1, min_row=2, max_row=counts_ws.max_row)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(labels)
    bar_chart.height = 8
    bar_chart.width = 18
    bar_chart.dLbls = DataLabelList()
    bar_chart.dLbls.showVal = True
    counts_ws.add_chart(bar_chart, "E2")

    pie_chart = PieChart()
    pie_chart.title = "2026 Q1 各领域文献占比"
    pie_data = Reference(counts_ws, min_col=2, min_row=1, max_row=counts_ws.max_row)
    pie_labels = Reference(counts_ws, min_col=1, min_row=2, max_row=counts_ws.max_row)
    pie_chart.add_data(pie_data, titles_from_data=True)
    pie_chart.set_categories(pie_labels)
    pie_chart.height = 10
    pie_chart.width = 14
    pie_chart.dLbls = DataLabelList()
    pie_chart.dLbls.showPercent = True
    pie_chart.dLbls.showLeaderLines = True
    counts_ws.add_chart(pie_chart, "E20")

    line_chart = LineChart()
    line_chart.title = "2026 Q1 月度文献总量趋势"
    line_chart.y_axis.title = "文献数"
    line_chart.x_axis.title = "月份"
    line_data = Reference(month_ws, min_col=2, min_row=1, max_row=month_ws.max_row)
    line_labels = Reference(month_ws, min_col=1, min_row=2, max_row=month_ws.max_row)
    line_chart.add_data(line_data, titles_from_data=True)
    line_chart.set_categories(line_labels)
    line_chart.height = 8
    line_chart.width = 15
    line_chart.dLbls = DataLabelList()
    line_chart.dLbls.showVal = True
    month_ws.add_chart(line_chart, "E2")

    stacked_bar = BarChart()
    stacked_bar.type = "bar"
    stacked_bar.grouping = "stacked"
    stacked_bar.overlap = 100
    stacked_bar.title = "各领域月度分布"
    stacked_bar.x_axis.title = "领域"
    stacked_bar.y_axis.title = "文献数"
    stacked_data = Reference(trend_ws, min_col=2, min_row=1, max_col=4, max_row=trend_ws.max_row)
    stacked_labels = Reference(trend_ws, min_col=1, min_row=2, max_row=trend_ws.max_row)
    stacked_bar.add_data(stacked_data, titles_from_data=True)
    stacked_bar.set_categories(stacked_labels)
    stacked_bar.height = 10
    stacked_bar.width = 18
    trend_ws.add_chart(stacked_bar, "G2")

    wb.save(OUTPUT_XLSX)


def build_markdown_report(category_counts: pd.DataFrame, monthly_trend: pd.DataFrame, month_total: pd.DataFrame) -> None:
    total = int(category_counts["count"].sum())
    top3 = category_counts.head(3)
    top3_text = "、".join(f"{row.category}（{row.count}篇）" for row in top3.itertuples(index=False))
    lines = [
        "# 2026 CCUS Q1 文献分析简报",
        "",
        "## 统计范围",
        "- 数据源：`all_literature_reclassified_11cats.xlsx` 与 `all_literature_classified.xlsx`",
        "- 统计口径：排除“其他”分类，仅统计 2026 年 1-3 月样本",
        f"- Q1 文献总量：{total} 篇",
        "",
        "## 核心结论",
    ]
    for insight in generate_insights(category_counts, monthly_trend, month_total):
        lines.append(f"- {insight}")

    lines.extend(
        [
            "",
            "## 各领域文献数",
            "| 排名 | 分类 | 文献数 | 占比 |",
            "| --- | --- | ---: | ---: |",
        ]
    )
    for idx, row in enumerate(category_counts.itertuples(index=False), start=1):
        lines.append(f"| {idx} | {row.category} | {row.count} | {row.share:.1%} |")

    lines.extend(
        [
            "",
            "## 月度总量趋势",
            "| 月份 | 文献数 | 占比 |",
            "| --- | ---: | ---: |",
        ]
    )
    for row in month_total.itertuples(index=False):
        lines.append(f"| {row.month} | {row.count} | {row.share:.1%} |")

    lines.extend(
        [
            "",
            "## 趋势概览",
            f"- 头部领域：{top3_text}",
            "- 结构判断：转化端研究继续主导，尤其是电转化和光转化；捕集与分离保持第二梯队核心位置。",
            "- 节奏判断：1 月样本最多，随后逐月回落；政策与产业化在 2 月相对集中，显示应用和治理议题阶段性升温。",
            "- 风险提示：该趋势反映当前样本集合，不宜直接外推为全学科完整发表趋势。",
        ]
    )

    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    q1 = load_q1_dataset()
    category_counts, monthly_trend, month_total = build_summary_tables(q1)
    build_excel_report(q1, category_counts, monthly_trend, month_total)
    build_markdown_report(category_counts, monthly_trend, month_total)
    print(f"Generated: {OUTPUT_XLSX.name}")
    print(f"Generated: {OUTPUT_MD.name}")


if __name__ == "__main__":
    main()
